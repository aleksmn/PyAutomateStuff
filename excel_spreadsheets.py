import openpyxl
import os

# os.getcwd()
# os.chdir() # change working directory
# os.path.abspath('.')
# os.listdir()
# os.path.getsize('excel/example.xlsx')

wb = openpyxl.load_workbook('excel/example.xlsx')
# wb.sheetnames
sheet = wb['Sheet1']

# sheet['A1'].value # datetime.datetime(2015, 4, 5, 13, 34, 2)
# sheet['B1'].value # 'Apples'

for i in range(1, 8):
    a = sheet.cell(row=i, column=1).value
    b = sheet.cell(row=i, column=2).value
    print(a, b)
