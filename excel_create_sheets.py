import openpyxl

wb = openpyxl.Workbook()
sheet = wb['Sheet']

sheet['A1'] = 'Hi'
sheet['B1'] = 'There!'
sheet['C1'] = 24

sheet2 = wb.create_sheet()

# wb.sheetnames

sheet2.title = 'My New Sheet Title'

for i in range(1, 8):
    for j in range(1, 5):
        sheet2.cell(row = i, column = j).value = i*j

wb.create_sheet(index=0, title="My First Worksheet")

wb.save('excel/new_example.xlsx')
